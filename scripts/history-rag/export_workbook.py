#!/usr/bin/env python3
"""Export the audited History RAG workbook to a deterministic NDJSON package."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook


PACKAGE_VERSION = "v1"
DEFAULT_OUTPUT_DIR = Path("data/history-rag/v1")
OUTPUT_FILES = (
    "historical-events.ndjson",
    "textbook-references.ndjson",
    "textbook-reference-removals.ndjson",
    "textbook-contents.ndjson",
    "textbook-content-refs.ndjson",
    "research-sources.ndjson",
    "event-research-sources.ndjson",
    "event-external-sources.ndjson",
)
EXPECTED_COUNTS = {
    "historicalEvents": 361,
    "textbookReferences": 386,
    "textbookReferenceRemovals": 9,
    "visibleTextbookReferences": 359,
    "hiddenTextbookReferences": 27,
    "activeTextbookEvents": 345,
    "textbookContents": 361,
    "researchSources": 231,
    "eventResearchSources": 1265,
    "eventExternalSources": 648,
    "exactExcerptPages": 13,
    "referenceRanges": 373,
    "internalLocalSources": 28,
}

EXPECTED_REMOVAL_IDS = {
    "REMOVE_WRONG_MAPPING": {120268, 120270, 120271, 120337, 120437, 120594},
    "REMOVE_QUARANTINED": {120303, 120327, 120609},
}


class ExportValidationError(RuntimeError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_text(value: str) -> str:
    return sha256_bytes(value.encode("utf-8"))


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True)


def clean_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def read_sheet(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ExportValidationError(f"Missing workbook sheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    if not headers or any(not header for header in headers):
        raise ExportValidationError(f"Invalid header row in sheet: {sheet_name}")

    result: list[dict[str, Any]] = []
    for values in rows:
        cleaned = [clean_cell(value) for value in values[: len(headers)]]
        if not any(value is not None for value in cleaned):
            continue
        result.append(dict(zip(headers, cleaned, strict=True)))
    return result


def require_unique(rows: Iterable[dict[str, Any]], key: str, label: str) -> None:
    values = [row.get(key) for row in rows]
    duplicates = sorted(value for value, count in Counter(values).items() if count > 1)
    if duplicates:
        raise ExportValidationError(f"Duplicate {label}: {duplicates[:10]}")
    if any(value is None or value == "" for value in values):
        raise ExportValidationError(f"Blank {label}")


def require_count(name: str, actual: int) -> None:
    expected = EXPECTED_COUNTS[name]
    if actual != expected:
        raise ExportValidationError(f"{name}: expected {expected}, got {actual}")


def parse_ref_ids(value: Any) -> list[int]:
    if value is None or str(value).strip() == "":
        return []
    return [int(part.strip()) for part in str(value).split(",") if part.strip()]


def natural_import_key(value: str) -> tuple[str, int, str]:
    match = re.fullmatch(r"([^0-9]*)([0-9]+)", value)
    if match:
        return match.group(1), int(match.group(2)), value
    return value, -1, value


def normalize_uri(uri: str | None) -> str | None:
    if uri is None or not uri.strip():
        return None
    value = uri.strip()
    if value.startswith("local:"):
        return value
    parts = urlsplit(value)
    if not parts.scheme or not parts.netloc:
        return value
    path = parts.path
    if path == "/":
        path = ""
    elif path.endswith("/"):
        path = path[:-1]
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), path, parts.query, parts.fragment))


def source_dedupe_key(
    source_type: str,
    import_key: str | None,
    canonical_uri: str | None,
    external_id: str | None,
) -> str:
    if external_id:
        identity_type, identity = "external_id", external_id.strip()
    elif canonical_uri:
        identity_type, identity = "canonical_uri", normalize_uri(canonical_uri)
    elif import_key:
        identity_type, identity = "import_key", import_key.strip()
    else:
        raise ExportValidationError("Source has no canonical dedupe identity")
    return sha256_text(canonical_json([source_type.lower(), identity_type, identity]))


def export_historical(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    require_count("historicalEvents", len(rows))
    require_unique(rows, "event_id", "historical event_id")
    fields = (
        "event_id",
        "title",
        "card_summary",
        "canonical_summary",
        "detailed_narrative",
        "significance",
    )
    return [{field: row.get(field) for field in fields} for row in sorted(rows, key=lambda item: item["event_id"])]


def export_textbook_references(
    rows: list[dict[str, Any]], audit_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    require_count("textbookReferences", len(rows))
    require_unique(rows, "id", "textbook reference id")
    require_unique(audit_rows, "id", "page audit reference id")
    audit_by_id = {int(row["id"]): row for row in audit_rows}
    active_ids = {int(row["id"]) for row in rows}
    missing_audit_ids = active_ids - set(audit_by_id)
    if missing_audit_ids:
        raise ExportValidationError(
            f"Full page audit is missing active textbook reference IDs: {sorted(missing_audit_ids)}"
        )

    exported: list[dict[str, Any]] = []
    for row in sorted(rows, key=lambda item: int(item["id"])):
        reference_id = int(row["id"])
        audit = audit_by_id[reference_id]
        page_scope = str(audit["page_scope"])
        page_basis = str(audit["page_number_basis"])
        if page_scope not in {"EXACT_EXCERPT_PAGE", "REFERENCE_RANGE"}:
            raise ExportValidationError(f"Invalid page_scope for ref {reference_id}: {page_scope}")
        if page_basis != "PRINTED_BOOK_PAGE":
            raise ExportValidationError(f"Invalid page_number_basis for ref {reference_id}: {page_basis}")
        if row["page_start"] != audit["final_page_start"] or row["page_end"] != audit["final_page_end"]:
            raise ExportValidationError(f"Page audit mismatch for ref {reference_id}")
        if row["page_start"] is None or row["page_end"] is None:
            raise ExportValidationError(f"Blank final page range for ref {reference_id}")
        if int(row["page_end"]) < int(row["page_start"]):
            raise ExportValidationError(f"Invalid final page range for ref {reference_id}")
        excerpt = row.get("excerpt")
        if not isinstance(excerpt, str) or not excerpt:
            raise ExportValidationError(f"Blank excerpt for ref {reference_id}")
        if sha256_text(excerpt) != audit.get("excerpt_after_hash"):
            raise ExportValidationError(f"Excerpt hash mismatch for ref {reference_id}")
        show_on_detail = row.get("show_on_detail")
        if show_on_detail not in {0, 1}:
            raise ExportValidationError(
                f"Invalid show_on_detail for ref {reference_id}: {show_on_detail}"
            )

        exported.append(
            {
                "id": reference_id,
                "event_id": row["event_id"],
                "grade": int(row["grade"]),
                "book": row["book"],
                "theme": row.get("theme"),
                "lesson": row.get("lesson"),
                "page_start": int(row["page_start"]),
                "page_end": int(row["page_end"]),
                "page_scope": page_scope,
                "page_number_basis": page_basis,
                "page_mapping_status": (
                    "EXACT_PAGE_MAPPED"
                    if page_scope == "EXACT_EXCERPT_PAGE"
                    else "REFERENCE_RANGE_MAPPED"
                ),
                "excerpt": excerpt,
                "url": row.get("url"),
                "source_key": row.get("source_key"),
                "show_on_detail": bool(show_on_detail),
            }
        )

    scopes = Counter(row["page_scope"] for row in exported)
    require_count("exactExcerptPages", scopes["EXACT_EXCERPT_PAGE"])
    require_count("referenceRanges", scopes["REFERENCE_RANGE"])
    require_count("visibleTextbookReferences", sum(row["show_on_detail"] for row in exported))
    require_count("hiddenTextbookReferences", sum(not row["show_on_detail"] for row in exported))
    require_count("activeTextbookEvents", len({row["event_id"] for row in exported}))
    return exported


def export_textbook_reference_removals(
    wrong_rows: list[dict[str, Any]], quarantined_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    categorized = [
        ("REMOVE_WRONG_MAPPING", row) for row in wrong_rows
    ] + [
        ("REMOVE_QUARANTINED", row) for row in quarantined_rows
    ]
    require_count("textbookReferenceRemovals", len(categorized))
    require_unique((row for _, row in categorized), "id", "textbook reference removal id")

    actual_by_category: defaultdict[str, set[int]] = defaultdict(set)
    result: list[dict[str, Any]] = []
    for category, row in categorized:
        reference_id = int(row["id"])
        actual_by_category[category].add(reference_id)
        result.append({
            "id": reference_id,
            "event_id": row["event_id"],
            "grade": int(row["grade"]),
            "book": row["book"],
            "theme": row.get("theme"),
            "lesson": row.get("lesson"),
            "page_start": int(row["page_start"]),
            "page_end": int(row["page_end"]),
            "excerpt": row["excerpt"],
            "url": row["url"],
            "source_key": row["source_key"],
            "created_at": row.get("created_at"),
            "removal_category": category,
            "semantic_status": row["semantic_status"],
            "audit_reason": row["audit_reason"],
            "recommended_action": row["recommended_action"],
        })

    for category, expected_ids in EXPECTED_REMOVAL_IDS.items():
        if actual_by_category[category] != expected_ids:
            raise ExportValidationError(
                f"{category} IDs: expected {sorted(expected_ids)}, got {sorted(actual_by_category[category])}"
            )
    return sorted(result, key=lambda item: item["id"])


def export_textbook_contents(
    rows: list[dict[str, Any]], references: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    require_count("textbookContents", len(rows))
    require_unique(rows, "event_id", "textbook content event_id")
    reference_by_id = {row["id"]: row for row in references}
    contents: list[dict[str, Any]] = []
    relations: list[dict[str, Any]] = []

    for row in sorted(rows, key=lambda item: item["event_id"]):
        event_id = row["event_id"]
        content = row.get("content")
        content_hash = row.get("content_hash")
        if content is not None and sha256_text(content) != content_hash:
            raise ExportValidationError(f"Textbook content hash mismatch for event {event_id}")
        if content is None and content_hash is not None:
            raise ExportValidationError(f"Null content has a hash for event {event_id}")

        ref_ids = parse_ref_ids(row.get("source_ref_ids"))
        if int(row.get("reference_count") or 0) != len(ref_ids):
            raise ExportValidationError(f"reference_count mismatch for event {event_id}")
        for source_order, reference_id in enumerate(ref_ids, start=1):
            reference = reference_by_id.get(reference_id)
            if reference is None or reference["event_id"] != event_id:
                raise ExportValidationError(
                    f"Invalid textbook content reference {reference_id} for event {event_id}"
                )
            relations.append(
                {
                    "event_id": event_id,
                    "textbook_ref_id": reference_id,
                    "source_order": source_order,
                }
            )

        contents.append(
            {
                "event_id": event_id,
                "content": content,
                "content_status": row["content_status"],
                "content_source": row["content_source"],
                "reference_count": int(row.get("reference_count") or 0),
                "grade_scope": row.get("grade_scope"),
                "correction_note": row.get("correction_note"),
                "content_hash": content_hash,
                "verified_at": row.get("verified_at"),
                "verified_by": row.get("verified_by"),
            }
        )
    return contents, relations


def export_research_sources(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    require_count("researchSources", len(rows))
    require_unique(rows, "source_id", "research source_id")
    result: list[dict[str, Any]] = []
    for row in sorted(rows, key=lambda item: natural_import_key(str(item["source_id"]))):
        import_key = str(row["source_id"])
        uri = normalize_uri(row.get("url"))
        internal = bool(uri and uri.startswith("local:"))
        source_type = "local" if internal else "research"
        result.append(
            {
                "import_key": import_key,
                "dedupe_key": source_dedupe_key(source_type, import_key, uri, None),
                "source_type": source_type,
                "title": row["source_name"],
                "canonical_uri": uri,
                "external_id": None,
                "language": None,
                "is_internal": internal,
                "source_role": row.get("source_role"),
                "usage_note": row.get("usage_in_batch"),
                "batch": row.get("batch"),
            }
        )
    require_count("internalLocalSources", sum(1 for row in result if row["is_internal"]))
    return result


def export_event_research_sources(
    rows: list[dict[str, Any]], source_keys: set[str], event_ids: set[str]
) -> list[dict[str, Any]]:
    require_count("eventResearchSources", len(rows))
    result: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        event_id = row["event_id"]
        import_key = str(row["source_id"])
        relation_key = (event_id, import_key)
        if event_id not in event_ids:
            raise ExportValidationError(f"Research mapping orphan event: {event_id}")
        if import_key not in source_keys:
            raise ExportValidationError(f"Research mapping missing source: {import_key}")
        if relation_key in seen:
            raise ExportValidationError(f"Duplicate research mapping: {relation_key}")
        seen.add(relation_key)
        result.append(
            {
                "event_id": event_id,
                "source_import_key": import_key,
                "source_order": int(row["source_order"]),
                "source_role": row.get("source_role"),
                "usage_note": row.get("usage_in_batch"),
                "verification_status": row.get("mapping_status"),
            }
        )
    return sorted(result, key=lambda item: (item["event_id"], item["source_order"], item["source_import_key"]))


def export_event_external_sources(
    rows: list[dict[str, Any]], event_ids: set[str]
) -> list[dict[str, Any]]:
    require_count("eventExternalSources", len(rows))
    result: list[dict[str, Any]] = []
    seen_pairs: set[tuple[str, str]] = set()
    order_by_event: defaultdict[str, int] = defaultdict(int)
    for row in rows:
        event_id = row["event_id"]
        if event_id not in event_ids:
            raise ExportValidationError(f"External source orphan event: {event_id}")
        canonical_uri = normalize_uri(row.get("url"))
        if canonical_uri is None:
            raise ExportValidationError(f"External source URL is blank for event {event_id}")
        pair = (event_id, canonical_uri)
        if pair in seen_pairs:
            raise ExportValidationError(f"Duplicate external event/URL pair: {pair}")
        seen_pairs.add(pair)
        order_by_event[event_id] += 1
        source_type = str(row["source_type"]).lower()
        external_id = str(row["external_id"]).strip() if row.get("external_id") else None
        dedupe_key = source_dedupe_key(source_type, None, canonical_uri, external_id)
        result.append(
            {
                "event_id": event_id,
                "source_order": order_by_event[event_id],
                "source_import_key": f"external:{dedupe_key}",
                "dedupe_key": dedupe_key,
                "source_type": source_type,
                "title": row["title"],
                "canonical_uri": canonical_uri,
                "external_id": external_id,
                "language": row.get("language"),
                "is_internal": canonical_uri.startswith("local:"),
                "match_type": row["match_type"],
                "is_primary": bool(row.get("is_primary")),
                "verification_status": row["verification_status"],
                "notes": row.get("notes"),
            }
        )
    return sorted(result, key=lambda item: (item["event_id"], item["source_order"], item["dedupe_key"]))


def write_ndjson(path: Path, rows: list[dict[str, Any]]) -> str:
    content = "".join(canonical_json(row) + "\n" for row in rows)
    path.write_text(content, encoding="utf-8", newline="\n")
    return sha256_bytes(content.encode("utf-8"))


def package_hash(file_hashes: dict[str, str]) -> str:
    payload = "".join(f"{name}\0{file_hashes[name]}\n" for name in OUTPUT_FILES)
    return sha256_text(payload)


def export_package(workbook_path: Path, output_dir: Path) -> dict[str, Any]:
    workbook_path = workbook_path.resolve()
    output_dir = output_dir.resolve()
    if not workbook_path.is_file():
        raise ExportValidationError(f"Workbook does not exist: {workbook_path}")

    workbook_sha256 = sha256_bytes(workbook_path.read_bytes())
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        historical = export_historical(read_sheet(workbook, "Historical Events Import"))
        textbook_references = export_textbook_references(
            read_sheet(workbook, "Event Textbook Refs Clean"),
            read_sheet(workbook, "Textbook Ref Full Page Audit"),
        )
        textbook_reference_removals = export_textbook_reference_removals(
            read_sheet(workbook, "Removed Textbook Refs"),
            read_sheet(workbook, "Textbook Ref Quarantine"),
        )
        active_reference_ids = {row["id"] for row in textbook_references}
        removed_reference_ids = {row["id"] for row in textbook_reference_removals}
        if active_reference_ids & removed_reference_ids:
            raise ExportValidationError("Active and removed textbook reference IDs overlap")
        if len(active_reference_ids | removed_reference_ids) != 395:
            raise ExportValidationError("Semantic textbook reference partition must contain 395 IDs")
        textbook_contents, textbook_content_refs = export_textbook_contents(
            read_sheet(workbook, "Event Textbook Contents"), textbook_references
        )
        research_sources = export_research_sources(read_sheet(workbook, "Research Sources Catalog"))
        event_ids = {row["event_id"] for row in historical}
        event_research_sources = export_event_research_sources(
            read_sheet(workbook, "Event Research Source Map"),
            {row["import_key"] for row in research_sources},
            event_ids,
        )
        event_external_sources = export_event_external_sources(
            read_sheet(workbook, "Event External Sources"), event_ids
        )
    finally:
        workbook.close()

    rows_by_file = {
        "historical-events.ndjson": historical,
        "textbook-references.ndjson": textbook_references,
        "textbook-reference-removals.ndjson": textbook_reference_removals,
        "textbook-contents.ndjson": textbook_contents,
        "textbook-content-refs.ndjson": textbook_content_refs,
        "research-sources.ndjson": research_sources,
        "event-research-sources.ndjson": event_research_sources,
        "event-external-sources.ndjson": event_external_sources,
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    file_hashes = {
        filename: write_ndjson(output_dir / filename, rows_by_file[filename])
        for filename in OUTPUT_FILES
    }
    counts = dict(EXPECTED_COUNTS)
    counts["textbookContentRefs"] = len(textbook_content_refs)
    manifest = {
        "packageVersion": PACKAGE_VERSION,
        "workbookSha256": workbook_sha256,
        "packageSha256": package_hash(file_hashes),
        "counts": counts,
        "files": file_hashes,
    }
    manifest_content = json.dumps(manifest, ensure_ascii=False, indent=2) + "\n"
    (output_dir / "manifest.json").write_text(manifest_content, encoding="utf-8", newline="\n")
    return manifest


def main() -> None:
    args = parse_args()
    manifest = export_package(args.workbook, args.output_dir)
    print(
        f"Exported History RAG package {manifest['packageVersion']} to "
        f"{args.output_dir.resolve()} (packageSha256={manifest['packageSha256']})"
    )


if __name__ == "__main__":
    main()
